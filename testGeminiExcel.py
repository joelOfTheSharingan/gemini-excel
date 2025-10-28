import google.generativeai as genai
import pandas as pd
import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import logging
import io

# --- Gemini Configuration ---
genai.configure(api_key="AIzaSyDyI7ihTGRkU8x5v3XoBtDv7nwzNW0vgGk")  # replace with env variable ideally
model = genai.GenerativeModel("gemini-2.0-flash")

# --- Logging Setup ---
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

# --- Color & Header Configurations ---
COLOR_MAP = {
    "Food": "FFEB9C",
    "Travel": "C6EFCE",
    "Bills": "FFCDD2",
    "Fuel": "D9E1F2",
    "Health": "FCE4D6",
    "Others": "D9D2E9",
}
SUMMARY_HEADER_FILL_COLOR = "E2EFDA"
TRANSACTION_HEADERS = ["Date", "Description", "Amount", "Category"]
SUMMARY_HEADERS = ["Category", "Total", "Percentage (%)"]
DEFAULT_FILL_COLOR = "FFFFFF"


def classify_transactions(text):
    """
    Uses Gemini to extract and classify transactions from a text statement.
    Returns a tuple: (data, needs_correction)
    """
    prompt = f"""
You are a financial statement parser.
Extract transactions as a valid JSON list. Each item must include:
- "date" (e.g., "12 Jul")
- "description"
- "amount" (numeric string)
- "category" (Food, Travel, Bills, Fuel, Health, Others)
Missing fields → "Missing". Output pure JSON only.

Input:
{text}
"""
    needs_correction = False
    try:
        response = model.generate_content(prompt)
        logging.info(f"[DEBUG] Raw Gemini response:\n{response.text}")

        # Remove any markdown formatting Gemini might include
        json_text = re.sub(r'```(?:json)?|```', '', response.text).strip()
        data = json.loads(json_text)

        if not isinstance(data, list):
            raise ValueError("Expected a JSON list of transactions.")

        required_fields = ["date", "description", "amount", "category"]
        for i, tx in enumerate(data):
            if not isinstance(tx, dict):
                logging.warning(f"Skipping malformed transaction at index {i}: {tx}")
                continue

            for field in required_fields:
                if field not in tx or not tx[field] or tx[field] == "Missing":
                    tx[field] = None
                    needs_correction = True

            if tx.get("amount"):
                try:
                    tx["amount"] = str(float(tx["amount"]))
                except ValueError:
                    tx["amount"] = None
                    needs_correction = True

        logging.info("Parsed transactions successfully.")
        return data, needs_correction

    except json.JSONDecodeError as e:
        logging.error(f"Gemini returned invalid JSON: {e}")
        raise
    except Exception as e:
        logging.error(f"Error during classification: {e}", exc_info=True)
        raise
def write_to_excel(data: list[dict], color_map: dict) -> bytes:
    """
    Writes categorized transaction data to an Excel workbook in memory.
    Returns the Excel file as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Alignment, Font, NamedStyle
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    from datetime import datetime
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Categorized Transactions"
    ws.freeze_panes = "A2"

    # Header Style
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True)
    header_style.alignment = Alignment(horizontal="center")

    headers = ["Date", "Description", "Amount", "Category"]
    ws.append(headers)
    for cell in ws[1]:
        cell.style = header_style

    category_totals = defaultdict(float)
    grand_total = 0.0

    # Sort and clean data
    clean_data = [tx for tx in data if all(tx.get(k) for k in ["date", "description", "amount", "category"])]

    try:
        current_year = datetime.now().year
        clean_data.sort(key=lambda x: datetime.strptime(f"{x['date']} {current_year}", "%d %b %Y"))
    except Exception:
        pass

    # Write data rows
    for tx in clean_data:
        ws.append([tx["date"], tx["description"], tx["amount"], tx["category"]])
        fill_color = color_map.get(tx["category"], "FFFFFF")
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        try:
            amt = float(tx["amount"])
            category_totals[tx["category"]] += amt
            grand_total += amt
        except Exception:
            continue

    # Summary Section
    ws.append([])
    ws.append(["Summary"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="center")

    summary_headers = ["Category", "Total", "Percentage (%)"]
    ws.append(summary_headers)
    for cell in ws[ws.max_row]:
        cell.style = header_style
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for cat, total in sorted(category_totals.items()):
        percent = round((total / grand_total * 100), 2) if grand_total else 0
        ws.append([cat, round(total, 2), percent])
        for c in range(1, len(summary_headers) + 1):
            ws.cell(row=ws.max_row, column=c).alignment = Alignment(horizontal="center")

    # Auto-fit columns
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(len(str(cell.value or "")) for cell in ws[col_letter])
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save to bytes
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.read()
